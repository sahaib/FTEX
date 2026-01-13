#!/usr/bin/env python3
"""
FTEX Analyzer v6.0
==================
Professional analysis with beautiful Excel and PDF reports.

Usage:
    python3 analyze.py                          # Analyze output/tickets.json
    python3 analyze.py --input data/tickets.json
    python3 analyze.py --no-ai                  # Statistical only
"""

import argparse
import json
import sys
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

# Rich imports
try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
    from rich import box
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False
    print("⚠️  Install 'rich' for beautiful terminal UI: pip install rich")

# Add parent to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))

# Import analysis engine
try:
    from shared.smart_detection import AnalysisEngine, UserConfig, OllamaClient, get_zombie_stats
except ImportError:
    try:
        from src.shared.smart_detection import AnalysisEngine, UserConfig, OllamaClient, get_zombie_stats
    except ImportError:
        from smart_detection import AnalysisEngine, UserConfig, OllamaClient, get_zombie_stats

# Excel/DataFrame
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# PDF with reportlab
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle, PageBreak, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Fallback PDF
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False


# =============================================================================
# COLOR PALETTE (Professional)
# =============================================================================
class Colors:
    """Professional color palette for reports."""
    # Primary
    PRIMARY = "1F4E79"           # Dark blue
    PRIMARY_LIGHT = "2E75B6"     # Medium blue
    PRIMARY_LIGHTER = "BDD7EE"   # Light blue
    
    # Headers
    HEADER_BG = "1F4E79"         # Dark blue
    HEADER_TEXT = "FFFFFF"       # White
    
    # Rows
    ROW_ALT = "F2F2F2"           # Light gray alternating
    ROW_WHITE = "FFFFFF"         # White
    
    # Status colors
    SUCCESS = "70AD47"           # Green
    WARNING = "FFC000"           # Yellow/Orange
    DANGER = "FF5050"            # Red
    INFO = "5B9BD5"              # Blue
    
    # Confidence
    HIGH_CONF = "C6EFCE"         # Light green bg
    HIGH_CONF_TEXT = "006100"    # Dark green text
    MED_CONF = "FFEB9C"          # Light yellow bg
    MED_CONF_TEXT = "9C5700"     # Dark yellow text
    LOW_CONF = "FFC7CE"          # Light red bg
    LOW_CONF_TEXT = "9C0006"     # Dark red text


# =============================================================================
# TERMINAL UI
# =============================================================================
class TerminalUI:
    """Rich terminal interface."""
    
    def __init__(self):
        self.console = Console() if RICH_AVAILABLE else None
        self.start_time = datetime.now()
    
    def print_header(self):
        if self.console:
            self.console.print(Panel.fit(
                "[bold cyan]FTEX Deep Analyzer[/bold cyan] v6.0\n"
                "[dim]Self-Validating AI Analysis[/dim]",
                border_style="cyan"
            ))
            self.console.print()
        else:
            print("\n" + "="*60)
            print("FTEX Deep Analyzer v6.0")
            print("="*60 + "\n")
    
    def print_config(self, ai_available: bool, model: str = None):
        if self.console:
            config_table = Table(show_header=False, box=box.SIMPLE, padding=(0, 2))
            config_table.add_column("Key", style="cyan")
            config_table.add_column("Value")
            
            config_table.add_row("Product", UserConfig.PRODUCT_NAME)
            config_table.add_row("Entity Type", UserConfig.ENTITY_NAME)
            config_table.add_row("Known Solutions", str(len(UserConfig.KNOWN_SOLUTIONS)))
            
            if ai_available:
                config_table.add_row("AI Status", f"[green]✓ Connected[/green] ({model})")
            else:
                config_table.add_row("AI Status", "[yellow]⚠ Fallback mode[/yellow]")
            
            self.console.print(Panel(config_table, title="Configuration", border_style="blue"))
            self.console.print()
        else:
            print(f"Product: {UserConfig.PRODUCT_NAME}")
            print(f"Entity: {UserConfig.ENTITY_NAME}")
            print(f"AI: {'Available' if ai_available else 'Fallback mode'}")
    
    def print_loaded(self, count: int, path: str):
        if self.console:
            self.console.print(f"📁 Loaded [bold]{count:,}[/bold] tickets from [cyan]{path}[/cyan]")
        else:
            print(f"Loaded {count:,} tickets from {path}")
    
    def create_progress(self):
        if self.console:
            return Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(complete_style="cyan", finished_style="green"),
                TaskProgressColumn(),
                console=self.console
            )
        return None
    
    def print_results(self, results: Dict):
        if self.console:
            meta = results.get('metadata', {})
            zombies = results.get('zombies', {})
            sla = results.get('sla', {})
            
            # Summary panel
            summary_table = Table(show_header=False, box=box.SIMPLE, padding=(0, 2))
            summary_table.add_column("", style="cyan")
            summary_table.add_column("", justify="right")
            
            summary_table.add_row("Total Tickets", f"[bold]{meta.get('total_tickets', 0):,}[/bold]")
            summary_table.add_row("Categories", str(len(results.get('categories', {}))))
            summary_table.add_row("Anomalies", str(len(results.get('anomalies', []))))
            
            zombie_count = zombies.get('true_zombie_count', 0)
            zombie_rate = zombies.get('zombie_rate', 0)
            zombie_style = "red" if zombie_rate > 10 else "yellow" if zombie_rate > 5 else "green"
            summary_table.add_row("True Zombies", f"[{zombie_style}]{zombie_count:,} ({zombie_rate}%)[/{zombie_style}]")
            
            self.console.print()
            self.console.print(Panel(summary_table, title="📊 Analysis Results", border_style="green"))
            
            # SLA panel
            frt = sla.get('first_response', {})
            res = sla.get('resolution', {})
            
            if frt or res:
                sla_table = Table(show_header=True, box=box.SIMPLE)
                sla_table.add_column("Metric")
                sla_table.add_column("Average", justify="right")
                sla_table.add_column("Compliance", justify="right")
                
                frt_comp = frt.get('compliance_rate', 0)
                frt_style = "green" if frt_comp >= 90 else "yellow" if frt_comp >= 70 else "red"
                sla_table.add_row(
                    "First Response",
                    f"{frt.get('avg', 'N/A')} hrs",
                    f"[{frt_style}]{frt_comp}%[/{frt_style}]"
                )
                
                res_comp = res.get('compliance_rate', 0)
                res_style = "green" if res_comp >= 90 else "yellow" if res_comp >= 70 else "red"
                sla_table.add_row(
                    "Resolution",
                    f"{res.get('avg', 'N/A')} hrs",
                    f"[{res_style}]{res_comp}%[/{res_style}]"
                )
                
                self.console.print(Panel(sla_table, title="⏱️ SLA Performance", border_style="blue"))
        else:
            print("\nRESULTS:")
            meta = results.get('metadata', {})
            zombies = results.get('zombies', {})
            print(f"  Tickets: {meta.get('total_tickets', 0):,}")
            print(f"  Zombies: {zombies.get('true_zombie_count', 0)} ({zombies.get('zombie_rate', 0)}%)")
    
    def print_outputs(self, outputs: Dict[str, str]):
        if self.console:
            self.console.print()
            self.console.print(Panel(
                "\n".join([f"  📄 [cyan]{name}[/cyan]: {path}" for name, path in outputs.items()]),
                title="📁 Generated Files",
                border_style="green",
            ))
        else:
            print("\nGenerated Files:")
            for name, path in outputs.items():
                print(f"  {name}: {path}")
    
    def print_completion(self):
        elapsed = (datetime.now() - self.start_time).total_seconds()
        if self.console:
            self.console.print()
            self.console.print(Panel(f"[bold green]✓ Analysis complete[/bold green] in {elapsed:.1f}s", border_style="green"))
        else:
            print(f"\n✓ Complete in {elapsed:.1f}s")
    
    def print_error(self, message: str):
        if self.console:
            self.console.print(f"[bold red]✗[/bold red] {message}")
        else:
            print(f"ERROR: {message}")


# =============================================================================
# PROFESSIONAL REPORT GENERATOR
# =============================================================================
class ReportGenerator:
    """Generate professional Excel, Markdown, and PDF reports."""
    
    def __init__(self, results: Dict, output_dir: Path):
        self.results = results
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.generated_date = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    def generate_all(self) -> Dict[str, str]:
        outputs = {}
        
        if PANDAS_AVAILABLE:
            excel_path = self.generate_excel()
            if excel_path:
                outputs['Excel Report'] = str(excel_path)
        
        md_path = self.generate_markdown()
        outputs['Markdown Summary'] = str(md_path)
        
        pdf_path = self.generate_pdf()
        if pdf_path:
            outputs['PDF Report'] = str(pdf_path)
        
        json_path = self.generate_json()
        outputs['Raw Data'] = str(json_path)
        
        return outputs
    
    # =========================================================================
    # EXCEL GENERATION (Professional)
    # =========================================================================
    def generate_excel(self) -> Optional[Path]:
        if not PANDAS_AVAILABLE:
            return None
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import FormulaRule, ColorScaleRule, DataBarRule
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            return None
        
        excel_path = self.output_dir / 'analysis_report.xlsx'
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color=Colors.HEADER_TEXT, size=11)
        header_fill = PatternFill(start_color=Colors.HEADER_BG, end_color=Colors.HEADER_BG, fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        alt_fill = PatternFill(start_color=Colors.ROW_ALT, end_color=Colors.ROW_ALT, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Confidence fills
        high_fill = PatternFill(start_color=Colors.HIGH_CONF, end_color=Colors.HIGH_CONF, fill_type="solid")
        med_fill = PatternFill(start_color=Colors.MED_CONF, end_color=Colors.MED_CONF, fill_type="solid")
        low_fill = PatternFill(start_color=Colors.LOW_CONF, end_color=Colors.LOW_CONF, fill_type="solid")
        
        # =====================================================================
        # SHEET 1: Executive Overview
        # =====================================================================
        ws = wb.active
        ws.title = "Executive Overview"
        
        meta = self.results.get('metadata', {})
        foundation = self.results.get('foundation', {})
        zombies = self.results.get('zombies', {})
        sla = self.results.get('sla', {})
        
        # Title
        ws['A1'] = "FTEX Analysis Report"
        ws['A1'].font = Font(bold=True, size=18, color=Colors.PRIMARY)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Generated: {self.generated_date}"
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # Key Metrics Section
        ws['A4'] = "KEY METRICS"
        ws['A4'].font = Font(bold=True, size=14, color=Colors.PRIMARY)
        
        metrics = [
            ("Total Tickets", meta.get('total_tickets', 0), ""),
            ("Date Range", f"{foundation.get('date_range', {}).get('earliest', 'N/A')[:10]} to {foundation.get('date_range', {}).get('latest', 'N/A')[:10]}", ""),
            ("Issue Categories", len(self.results.get('categories', {})), ""),
            (f"{UserConfig.ENTITY_NAME_PLURAL.title()}", self.results.get('entities', {}).get('total_entities', 0), ""),
            ("", "", ""),
            ("True Zombies", zombies.get('true_zombie_count', 0), f"{zombies.get('zombie_rate', 0)}%"),
            ("Anomalies Detected", len(self.results.get('anomalies', [])), ""),
            ("", "", ""),
            ("FRT Compliance", f"{sla.get('first_response', {}).get('compliance_rate', 'N/A')}%", f"Avg: {sla.get('first_response', {}).get('avg', 'N/A')} hrs"),
            ("Resolution Compliance", f"{sla.get('resolution', {}).get('compliance_rate', 'N/A')}%", f"Avg: {sla.get('resolution', {}).get('avg', 'N/A')} hrs"),
        ]
        
        for i, (label, value, extra) in enumerate(metrics, start=5):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True) if label else Font()
            ws[f'B{i}'] = value
            ws[f'B{i}'].alignment = Alignment(horizontal='right')
            ws[f'C{i}'] = extra
            ws[f'C{i}'].font = Font(color="666666")
            
            # Highlight zombie row
            if label == "True Zombies":
                zombie_rate = zombies.get('zombie_rate', 0)
                if zombie_rate > 10:
                    ws[f'B{i}'].fill = PatternFill(start_color=Colors.LOW_CONF, end_color=Colors.LOW_CONF, fill_type="solid")
                elif zombie_rate > 5:
                    ws[f'B{i}'].fill = PatternFill(start_color=Colors.MED_CONF, end_color=Colors.MED_CONF, fill_type="solid")
                else:
                    ws[f'B{i}'].fill = PatternFill(start_color=Colors.HIGH_CONF, end_color=Colors.HIGH_CONF, fill_type="solid")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        
        # =====================================================================
        # SHEET 2: Issue Categories
        # =====================================================================
        ws2 = wb.create_sheet("Issue Categories")
        categories = self.results.get('categories', {})
        
        headers = ['Category', 'Description', 'Tickets', 'Zombies', 'Zombie %', 'Avg Resolution (days)', 'Root Causes', 'Sample IDs']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for cat_name, cat_data in sorted(categories.items(), key=lambda x: x[1].get('total_tickets', 0), reverse=True):
            ws2.cell(row=row, column=1, value=cat_name).border = thin_border
            ws2.cell(row=row, column=2, value=cat_data.get('description', '')[:80]).border = thin_border
            ws2.cell(row=row, column=3, value=cat_data.get('total_tickets', 0)).border = thin_border
            ws2.cell(row=row, column=4, value=cat_data.get('zombie_count', 0)).border = thin_border
            
            zombie_pct = cat_data.get('zombie_rate', 0)
            cell = ws2.cell(row=row, column=5, value=zombie_pct)
            cell.border = thin_border
            cell.number_format = '0.0%' if zombie_pct <= 1 else '0.0'
            if zombie_pct > 15:
                cell.fill = low_fill
            elif zombie_pct > 8:
                cell.fill = med_fill
            
            ws2.cell(row=row, column=6, value=cat_data.get('avg_resolution_days', '')).border = thin_border
            ws2.cell(row=row, column=7, value=', '.join(cat_data.get('typical_root_causes', [])[:2])).border = thin_border
            ws2.cell(row=row, column=8, value=', '.join(map(str, cat_data.get('ticket_ids', [])[:5]))).border = thin_border
            
            # Alternating row colors
            if row % 2 == 0:
                for col in range(1, 9):
                    if not ws2.cell(row=row, column=col).fill.start_color.rgb or ws2.cell(row=row, column=col).fill.start_color.rgb == '00000000':
                        ws2.cell(row=row, column=col).fill = alt_fill
            row += 1
        
        # Auto-filter and freeze
        ws2.auto_filter.ref = f"A1:H{row-1}"
        ws2.freeze_panes = 'A2'
        
        # Column widths
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 18
        ws2.column_dimensions['G'].width = 35
        ws2.column_dimensions['H'].width = 25
        
        # =====================================================================
        # SHEET 3: Entities (Vessels/Stores/etc)
        # =====================================================================
        ws3 = wb.create_sheet(UserConfig.ENTITY_NAME_PLURAL.title())
        entities = self.results.get('entities', {}).get('entities', {})
        
        headers = [UserConfig.ENTITY_NAME.title(), 'Total Tickets', 'Zombies', 'Zombie %', 'Top Issue', 'Avg Resolution (days)']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for entity_name, entity_data in sorted(entities.items(), key=lambda x: x[1].get('total_tickets', 0), reverse=True)[:100]:
            ws3.cell(row=row, column=1, value=entity_name).border = thin_border
            ws3.cell(row=row, column=2, value=entity_data.get('total_tickets', 0)).border = thin_border
            ws3.cell(row=row, column=3, value=entity_data.get('zombie_count', 0)).border = thin_border
            
            zombie_pct = entity_data.get('zombie_rate', 0)
            cell = ws3.cell(row=row, column=4, value=zombie_pct)
            cell.border = thin_border
            if zombie_pct > 20:
                cell.fill = low_fill
            elif zombie_pct > 10:
                cell.fill = med_fill
            
            top_issue = entity_data.get('top_issue', ('', 0))
            ws3.cell(row=row, column=5, value=top_issue[0] if isinstance(top_issue, tuple) else str(top_issue)).border = thin_border
            ws3.cell(row=row, column=6, value=entity_data.get('avg_resolution_days', '')).border = thin_border
            
            if row % 2 == 0:
                for col in range(1, 7):
                    if not ws3.cell(row=row, column=col).fill.start_color.rgb or ws3.cell(row=row, column=col).fill.start_color.rgb == '00000000':
                        ws3.cell(row=row, column=col).fill = alt_fill
            row += 1
        
        ws3.auto_filter.ref = f"A1:F{row-1}"
        ws3.freeze_panes = 'A2'
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 10
        ws3.column_dimensions['D'].width = 10
        ws3.column_dimensions['E'].width = 30
        ws3.column_dimensions['F'].width = 18
        
        # =====================================================================
        # SHEET 4: Anomalies
        # =====================================================================
        ws4 = wb.create_sheet("Anomalies")
        anomalies = self.results.get('anomalies', [])
        
        headers = ['Type', 'Severity', 'Entity', 'Description', 'Ticket IDs']
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for anom in anomalies:
            ws4.cell(row=row, column=1, value=anom.get('type', '').replace('_', ' ').title()).border = thin_border
            
            severity = anom.get('severity', 'medium').lower()
            sev_cell = ws4.cell(row=row, column=2, value=severity.title())
            sev_cell.border = thin_border
            if severity == 'high':
                sev_cell.fill = low_fill
                sev_cell.font = Font(bold=True, color=Colors.LOW_CONF_TEXT)
            elif severity == 'medium':
                sev_cell.fill = med_fill
                sev_cell.font = Font(color=Colors.MED_CONF_TEXT)
            else:
                sev_cell.fill = high_fill
            
            ws4.cell(row=row, column=3, value=anom.get('entity', '')).border = thin_border
            ws4.cell(row=row, column=4, value=anom.get('description', '')).border = thin_border
            ws4.cell(row=row, column=5, value=', '.join(map(str, anom.get('ticket_ids', [])[:5]))).border = thin_border
            
            if row % 2 == 0:
                for col in range(1, 6):
                    if col != 2:  # Keep severity color
                        if not ws4.cell(row=row, column=col).fill.start_color.rgb or ws4.cell(row=row, column=col).fill.start_color.rgb == '00000000':
                            ws4.cell(row=row, column=col).fill = alt_fill
            row += 1
        
        ws4.auto_filter.ref = f"A1:E{row-1}"
        ws4.freeze_panes = 'A2'
        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 12
        ws4.column_dimensions['C'].width = 25
        ws4.column_dimensions['D'].width = 50
        ws4.column_dimensions['E'].width = 30
        
        # =====================================================================
        # SHEET 5: Zombie Tickets (Action Required)
        # =====================================================================
        ws5 = wb.create_sheet("Zombie Tickets")
        zombie_list = self.results.get('zombies', {}).get('true_zombies', [])
        
        # Warning header
        ws5['A1'] = "⚠️ ACTION REQUIRED: These tickets have NO customer response"
        ws5['A1'].font = Font(bold=True, size=12, color=Colors.LOW_CONF_TEXT)
        ws5['A1'].fill = PatternFill(start_color=Colors.LOW_CONF, end_color=Colors.LOW_CONF, fill_type="solid")
        ws5.merge_cells('A1:F1')
        
        headers = ['Ticket ID', 'Subject', 'Created', 'Status', 'Priority', 'Reason']
        for col, header in enumerate(headers, 1):
            cell = ws5.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 3
        for z in zombie_list[:200]:
            ws5.cell(row=row, column=1, value=z.get('ticket_id')).border = thin_border
            ws5.cell(row=row, column=2, value=z.get('subject', '')[:60]).border = thin_border
            ws5.cell(row=row, column=3, value=z.get('created_at', '')[:10] if z.get('created_at') else '').border = thin_border
            ws5.cell(row=row, column=4, value=z.get('status', '')).border = thin_border
            
            priority = z.get('priority', '')
            pri_cell = ws5.cell(row=row, column=5, value=priority)
            pri_cell.border = thin_border
            if priority == 'Urgent':
                pri_cell.fill = low_fill
                pri_cell.font = Font(bold=True)
            elif priority == 'High':
                pri_cell.fill = med_fill
            
            ws5.cell(row=row, column=6, value=z.get('reason', '')).border = thin_border
            
            if row % 2 == 1:
                for col in range(1, 7):
                    if col != 5 or not priority in ['Urgent', 'High']:
                        if not ws5.cell(row=row, column=col).fill.start_color.rgb or ws5.cell(row=row, column=col).fill.start_color.rgb == '00000000':
                            ws5.cell(row=row, column=col).fill = alt_fill
            row += 1
        
        ws5.auto_filter.ref = f"A2:F{row-1}"
        ws5.freeze_panes = 'A3'
        ws5.column_dimensions['A'].width = 12
        ws5.column_dimensions['B'].width = 45
        ws5.column_dimensions['C'].width = 12
        ws5.column_dimensions['D'].width = 12
        ws5.column_dimensions['E'].width = 10
        ws5.column_dimensions['F'].width = 25
        
        # =====================================================================
        # SHEET 6: SLA Performance
        # =====================================================================
        ws6 = wb.create_sheet("SLA Performance")
        sla_data = self.results.get('sla', {})
        
        ws6['A1'] = "SLA Performance Summary"
        ws6['A1'].font = Font(bold=True, size=14, color=Colors.PRIMARY)
        ws6.merge_cells('A1:E1')
        
        headers = ['Priority', 'FRT Target (hrs)', 'FRT Avg (hrs)', 'FRT Compliance %', 'Resolution Target (hrs)', 'Resolution Avg (hrs)', 'Resolution Compliance %']
        for col, header in enumerate(headers, 1):
            cell = ws6.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Overall row
        frt = sla_data.get('first_response', {})
        res = sla_data.get('resolution', {})
        
        row = 4
        ws6.cell(row=row, column=1, value="OVERALL").font = Font(bold=True)
        ws6.cell(row=row, column=1).border = thin_border
        ws6.cell(row=row, column=2, value="-").border = thin_border
        ws6.cell(row=row, column=3, value=frt.get('avg', '')).border = thin_border
        
        frt_comp = frt.get('compliance_rate', 0)
        frt_cell = ws6.cell(row=row, column=4, value=frt_comp)
        frt_cell.border = thin_border
        if frt_comp >= 90:
            frt_cell.fill = high_fill
        elif frt_comp >= 70:
            frt_cell.fill = med_fill
        else:
            frt_cell.fill = low_fill
        
        ws6.cell(row=row, column=5, value="-").border = thin_border
        ws6.cell(row=row, column=6, value=res.get('avg', '')).border = thin_border
        
        res_comp = res.get('compliance_rate', 0)
        res_cell = ws6.cell(row=row, column=7, value=res_comp)
        res_cell.border = thin_border
        if res_comp >= 90:
            res_cell.fill = high_fill
        elif res_comp >= 70:
            res_cell.fill = med_fill
        else:
            res_cell.fill = low_fill
        
        # By priority
        by_priority = sla_data.get('by_priority', {})
        frt_targets = {'Urgent': 1, 'High': 4, 'Medium': 8, 'Low': 24}
        res_targets = {'Urgent': 4, 'High': 24, 'Medium': 72, 'Low': 168}
        
        row = 5
        for priority in ['Urgent', 'High', 'Medium', 'Low']:
            p_data = by_priority.get(priority, {})
            p_frt = p_data.get('first_response', {})
            p_res = p_data.get('resolution', {})
            
            ws6.cell(row=row, column=1, value=priority).border = thin_border
            ws6.cell(row=row, column=2, value=frt_targets.get(priority, '')).border = thin_border
            ws6.cell(row=row, column=3, value=p_frt.get('avg', '')).border = thin_border
            
            p_frt_comp = p_frt.get('compliance_rate', 0)
            cell = ws6.cell(row=row, column=4, value=p_frt_comp)
            cell.border = thin_border
            if p_frt_comp >= 90:
                cell.fill = high_fill
            elif p_frt_comp >= 70:
                cell.fill = med_fill
            elif p_frt_comp > 0:
                cell.fill = low_fill
            
            ws6.cell(row=row, column=5, value=res_targets.get(priority, '')).border = thin_border
            ws6.cell(row=row, column=6, value=p_res.get('avg', '')).border = thin_border
            
            p_res_comp = p_res.get('compliance_rate', 0)
            cell = ws6.cell(row=row, column=7, value=p_res_comp)
            cell.border = thin_border
            if p_res_comp >= 90:
                cell.fill = high_fill
            elif p_res_comp >= 70:
                cell.fill = med_fill
            elif p_res_comp > 0:
                cell.fill = low_fill
            
            if row % 2 == 1:
                for col in range(1, 8):
                    if col not in [4, 7]:
                        ws6.cell(row=row, column=col).fill = alt_fill
            row += 1
        
        ws6.column_dimensions['A'].width = 12
        ws6.column_dimensions['B'].width = 16
        ws6.column_dimensions['C'].width = 14
        ws6.column_dimensions['D'].width = 18
        ws6.column_dimensions['E'].width = 20
        ws6.column_dimensions['F'].width = 18
        ws6.column_dimensions['G'].width = 22
        
        # =====================================================================
        # SHEET 7: Key Findings
        # =====================================================================
        ws7 = wb.create_sheet("Key Findings")
        findings = self.results.get('findings', [])
        
        ws7['A1'] = "Evidence-Based Findings"
        ws7['A1'].font = Font(bold=True, size=14, color=Colors.PRIMARY)
        ws7.merge_cells('A1:F1')
        
        headers = ['#', 'Finding', 'Confidence', 'Severity', 'Evidence', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws7.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 4
        for i, f in enumerate(findings, 1):
            ws7.cell(row=row, column=1, value=i).border = thin_border
            ws7.cell(row=row, column=2, value=f.get('title', '')).border = thin_border
            
            confidence = f.get('confidence', '').lower()
            conf_cell = ws7.cell(row=row, column=3, value=confidence.title())
            conf_cell.border = thin_border
            conf_cell.alignment = Alignment(horizontal='center')
            if confidence == 'high':
                conf_cell.fill = high_fill
                conf_cell.font = Font(bold=True, color=Colors.HIGH_CONF_TEXT)
            elif confidence == 'medium':
                conf_cell.fill = med_fill
                conf_cell.font = Font(color=Colors.MED_CONF_TEXT)
            else:
                conf_cell.fill = low_fill
                conf_cell.font = Font(color=Colors.LOW_CONF_TEXT)
            
            severity = f.get('severity', '').lower()
            sev_cell = ws7.cell(row=row, column=4, value=severity.title())
            sev_cell.border = thin_border
            sev_cell.alignment = Alignment(horizontal='center')
            if severity == 'high':
                sev_cell.fill = low_fill
            elif severity == 'medium':
                sev_cell.fill = med_fill
            
            ws7.cell(row=row, column=5, value=f"{f.get('evidence_count', 0)} tickets").border = thin_border
            ws7.cell(row=row, column=6, value=f.get('recommendation', '')).border = thin_border
            
            if row % 2 == 0:
                for col in range(1, 7):
                    if col not in [3, 4]:
                        ws7.cell(row=row, column=col).fill = alt_fill
            row += 1
        
        ws7.auto_filter.ref = f"A3:F{row-1}"
        ws7.freeze_panes = 'A4'
        ws7.column_dimensions['A'].width = 5
        ws7.column_dimensions['B'].width = 45
        ws7.column_dimensions['C'].width = 12
        ws7.column_dimensions['D'].width = 12
        ws7.column_dimensions['E'].width = 12
        ws7.column_dimensions['F'].width = 50
        
        # Save workbook
        wb.save(excel_path)
        return excel_path
    
    # =========================================================================
    # PDF GENERATION (Professional)
    # =========================================================================
    def generate_pdf(self) -> Optional[Path]:
        pdf_path = self.output_dir / 'analysis_report.pdf'
        
        if REPORTLAB_AVAILABLE:
            return self._generate_pdf_reportlab(pdf_path)
        elif WEASYPRINT_AVAILABLE:
            return self._generate_pdf_weasyprint(pdf_path)
        return None
    
    def _generate_pdf_reportlab(self, pdf_path: Path) -> Path:
        """Generate professional PDF with reportlab."""
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        # Custom styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#1F4E79'),
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1F4E79'),
            spaceBefore=20,
            spaceAfter=10
        )
        
        subheading_style = ParagraphStyle(
            'CustomSubheading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#2E75B6'),
            spaceBefore=15,
            spaceAfter=8
        )
        
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            spaceAfter=8
        )
        
        story = []
        
        # Title
        story.append(Paragraph("FTEX Analysis Report", title_style))
        story.append(Paragraph(f"<i>Generated: {self.generated_date}</i>", body_style))
        story.append(Paragraph(f"<i>Product: {UserConfig.PRODUCT_NAME}</i>", body_style))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        
        meta = self.results.get('metadata', {})
        foundation = self.results.get('foundation', {})
        zombies = self.results.get('zombies', {})
        sla = self.results.get('sla', {})
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Tickets', f"{meta.get('total_tickets', 0):,}"],
            ['Date Range', f"{foundation.get('date_range', {}).get('earliest', 'N/A')[:10]} to {foundation.get('date_range', {}).get('latest', 'N/A')[:10]}"],
            ['Issue Categories', str(len(self.results.get('categories', {})))],
            ['True Zombies', f"{zombies.get('true_zombie_count', 0):,} ({zombies.get('zombie_rate', 0)}%)"],
            ['Anomalies', str(len(self.results.get('anomalies', [])))],
        ]
        
        summary_table = RLTable(summary_data, colWidths=[2.5*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # SLA Performance
        story.append(Paragraph("SLA Performance", heading_style))
        
        frt = sla.get('first_response', {})
        res = sla.get('resolution', {})
        
        sla_data = [
            ['Metric', 'Average', 'Compliance'],
            ['First Response Time', f"{frt.get('avg', 'N/A')} hours", f"{frt.get('compliance_rate', 'N/A')}%"],
            ['Resolution Time', f"{res.get('avg', 'N/A')} hours", f"{res.get('compliance_rate', 'N/A')}%"],
        ]
        
        sla_table = RLTable(sla_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        sla_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(sla_table)
        story.append(Spacer(1, 20))
        
        # Key Findings
        findings = self.results.get('findings', [])
        if findings:
            story.append(Paragraph("Key Findings", heading_style))
            
            for i, f in enumerate(findings[:8], 1):
                conf = f.get('confidence', '').title()
                conf_color = '#006100' if conf == 'High' else '#9C5700' if conf == 'Medium' else '#9C0006'
                
                story.append(Paragraph(
                    f"<b>{i}. {f.get('title', 'Finding')}</b>",
                    subheading_style
                ))
                story.append(Paragraph(
                    f"<font color='{conf_color}'><b>Confidence:</b> {conf}</font> ({f.get('evidence_count', 0)} tickets)",
                    body_style
                ))
                if f.get('recommendation'):
                    story.append(Paragraph(
                        f"<b>Recommendation:</b> {f.get('recommendation')}",
                        body_style
                    ))
                story.append(Spacer(1, 10))
        
        # Top Issue Categories
        story.append(PageBreak())
        story.append(Paragraph("Top Issue Categories", heading_style))
        
        categories = self.results.get('categories', {})
        cat_data = [['Category', 'Tickets', 'Zombies', 'Zombie %']]
        for cat_name, cat_info in sorted(categories.items(), key=lambda x: x[1].get('total_tickets', 0), reverse=True)[:10]:
            cat_data.append([
                cat_name[:35],
                str(cat_info.get('total_tickets', 0)),
                str(cat_info.get('zombie_count', 0)),
                f"{cat_info.get('zombie_rate', 0)}%"
            ])
        
        if len(cat_data) > 1:
            cat_table = RLTable(cat_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(cat_table)
        
        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            "<i>Generated by FTEX Analyzer v6.0</i>",
            ParagraphStyle('Footer', parent=body_style, textColor=colors.gray, alignment=TA_CENTER)
        ))
        
        doc.build(story)
        return pdf_path
    
    def _generate_pdf_weasyprint(self, pdf_path: Path) -> Path:
        """Fallback PDF generation with weasyprint."""
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; }}
                h1 {{ color: #1F4E79; border-bottom: 2px solid #1F4E79; padding-bottom: 10px; }}
                h2 {{ color: #2E75B6; margin-top: 30px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
                th {{ background-color: #1F4E79; color: white; padding: 10px; text-align: left; }}
                td {{ border: 1px solid #ddd; padding: 8px; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .metric {{ font-size: 24px; color: #1F4E79; font-weight: bold; }}
                .label {{ color: #666; }}
            </style>
        </head>
        <body>
            <h1>FTEX Analysis Report</h1>
            <p><em>Generated: {self.generated_date}</em></p>
        """
        
        meta = self.results.get('metadata', {})
        zombies = self.results.get('zombies', {})
        
        html_content += f"""
            <h2>Executive Summary</h2>
            <table>
                <tr><th>Metric</th><th>Value</th></tr>
                <tr><td>Total Tickets</td><td>{meta.get('total_tickets', 0):,}</td></tr>
                <tr><td>True Zombies</td><td>{zombies.get('true_zombie_count', 0):,} ({zombies.get('zombie_rate', 0)}%)</td></tr>
                <tr><td>Issue Categories</td><td>{len(self.results.get('categories', {}))}</td></tr>
                <tr><td>Anomalies</td><td>{len(self.results.get('anomalies', []))}</td></tr>
            </table>
        """
        
        html_content += """
            <p style="margin-top: 50px; color: #999; text-align: center;">
                <em>Generated by FTEX Analyzer v6.0</em>
            </p>
        </body>
        </html>
        """
        
        HTML(string=html_content).write_pdf(str(pdf_path))
        return pdf_path
    
    # =========================================================================
    # MARKDOWN GENERATION
    # =========================================================================
    def generate_markdown(self) -> Path:
        md_path = self.output_dir / 'analysis_summary.md'
        
        meta = self.results.get('metadata', {})
        foundation = self.results.get('foundation', {})
        zombies = self.results.get('zombies', {})
        categories = self.results.get('categories', {})
        findings = self.results.get('findings', [])
        sla = self.results.get('sla', {})
        
        lines = [
            f"# FTEX Analysis Report",
            f"",
            f"**Generated:** {self.generated_date}",
            f"**Product:** {UserConfig.PRODUCT_NAME}",
            f"**AI Enabled:** {'Yes' if meta.get('ai_enabled') else 'No'}",
            f"",
            f"---",
            f"",
            f"## Executive Summary",
            f"",
            f"| Metric | Value |",
            f"|--------|-------|",
            f"| Total Tickets | {meta.get('total_tickets', 0):,} |",
            f"| Date Range | {foundation.get('date_range', {}).get('earliest', 'N/A')[:10]} to {foundation.get('date_range', {}).get('latest', 'N/A')[:10]} |",
            f"| **True Zombies** | **{zombies.get('true_zombie_count', 0):,}** ({zombies.get('zombie_rate', 0)}%) |",
            f"| Issue Categories | {len(categories)} |",
            f"| Anomalies | {len(self.results.get('anomalies', []))} |",
            f"",
        ]
        
        # Findings
        if findings:
            lines.extend([f"## Key Findings", f""])
            for i, f in enumerate(findings[:10], 1):
                conf = f.get('confidence', '').lower()
                conf_icon = "🟢" if conf == 'high' else "🟡" if conf == 'medium' else "🔴"
                lines.append(f"### {i}. {f.get('title', 'Finding')}")
                lines.append(f"")
                lines.append(f"**Confidence:** {conf_icon} {conf.title()} ({f.get('evidence_count', 0)} tickets)")
                lines.append(f"")
                if f.get('recommendation'):
                    lines.append(f"**Recommendation:** {f['recommendation']}")
                lines.append(f"")
        
        # Categories
        if categories:
            lines.extend([f"## Top Issue Categories", f"", f"| Category | Tickets | Zombies | Zombie % |", f"|----------|---------|---------|----------|"])
            for cat, data in sorted(categories.items(), key=lambda x: x[1].get('total_tickets', 0), reverse=True)[:10]:
                lines.append(f"| {cat} | {data.get('total_tickets', 0)} | {data.get('zombie_count', 0)} | {data.get('zombie_rate', 0)}% |")
            lines.append(f"")
        
        # SLA
        frt = sla.get('first_response', {})
        res = sla.get('resolution', {})
        lines.extend([
            f"## SLA Performance",
            f"",
            f"| Metric | Average | Compliance |",
            f"|--------|---------|------------|",
            f"| First Response | {frt.get('avg', 'N/A')} hrs | {frt.get('compliance_rate', 'N/A')}% |",
            f"| Resolution | {res.get('avg', 'N/A')} hrs | {res.get('compliance_rate', 'N/A')}% |",
            f"",
            f"---",
            f"*Generated by FTEX Analyzer v6.0*",
        ])
        
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return md_path
    
    # =========================================================================
    # JSON EXPORT
    # =========================================================================
    def generate_json(self) -> Path:
        json_path = self.output_dir / 'analysis_data.json'
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, default=str)
        
        return json_path


# =============================================================================
# MAIN ANALYSIS FUNCTION
# =============================================================================
def main(input_path: str = None, output_dir: str = None, use_ai: bool = True, 
         clear_cache: bool = False, progress_callback=None):
    """Run complete analysis pipeline."""
    
    ui = TerminalUI()
    ui.print_header()
    
    # Find input file
    if input_path:
        tickets_path = Path(input_path)
    else:
        for candidate in ['output/tickets.json', 'tickets.json', '../output/tickets.json']:
            if Path(candidate).exists():
                tickets_path = Path(candidate)
                break
        else:
            ui.print_error("No tickets.json found. Run extraction first.")
            return None
    
    if not tickets_path.exists():
        ui.print_error(f"File not found: {tickets_path}")
        return None
    
    # Load tickets
    with open(tickets_path, 'r', encoding='utf-8') as f:
        tickets = json.load(f)
    
    if isinstance(tickets, dict):
        tickets = tickets.get('tickets', [])
    
    ui.print_loaded(len(tickets), str(tickets_path))
    
    # Check AI
    ai_available = False
    model = None
    if use_ai:
        client = OllamaClient()
        ai_available = client.available  # Property, not method
        if ai_available:
            model = client.model
    
    ui.print_config(ai_available, model)
    
    # Clear cache if requested
    if clear_cache:
        cache_path = Path(output_dir or 'reports') / 'analysis_cache.json'
        if cache_path.exists():
            cache_path.unlink()
    
    # Run analysis
    engine = AnalysisEngine(tickets, use_ai=ai_available)
    
    progress = ui.create_progress()
    if progress:
        with progress:
            task = progress.add_task("Analyzing...", total=100)
            
            def update_progress(stage, pct):
                progress.update(task, completed=pct, description=f"{stage}...")
            
            results = engine.run_analysis(progress_callback=update_progress)
    else:
        results = engine.run_analysis(progress_callback=progress_callback)
    
    ui.print_results(results)
    
    # Generate reports
    report_dir = Path(output_dir) if output_dir else Path('reports')
    generator = ReportGenerator(results, report_dir)
    outputs = generator.generate_all()
    
    ui.print_outputs(outputs)
    ui.print_completion()
    
    return results


# =============================================================================
# CLI
# =============================================================================
def analyze_main(args):
    """Entry point for CLI."""
    return main(
        input_path=args.input,
        output_dir=args.output,
        use_ai=not args.no_ai,
        clear_cache=args.clear_cache
    )


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='FTEX Analyzer v6.0')
    parser.add_argument('--input', '-i', help='Input tickets.json path')
    parser.add_argument('--output', '-o', help='Output directory', default='reports')
    parser.add_argument('--no-ai', action='store_true', help='Disable AI analysis')
    parser.add_argument('--clear-cache', action='store_true', help='Clear category cache')
    
    args = parser.parse_args()
    main(
        input_path=args.input,
        output_dir=args.output,
        use_ai=not args.no_ai,
        clear_cache=args.clear_cache
    )