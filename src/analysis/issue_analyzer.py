#!/usr/bin/env python3
"""
FTEX Issue-Centric Analyzer v7.0
================================
Streaming, conversation-aware, issue-deduplication analyzer.

Handles:
- 1.5GB+ files via ijson streaming
- Multi-issue ticket detection (new issues raised mid-conversation)
- Issue deduplication across tickets (same root issue → one cluster)
- High-precision new issue detection

Outputs:
- A) Top 50 issues with ticket lists (executive summary)
- B) Every distinct issue mapped to tickets (complete registry)
- C) Actionable issues only (filtered by severity/frequency)

Usage:
    python3 issue_analyzer.py --input tickets.json
    python3 issue_analyzer.py --input tickets.json --sample 1000
"""

import argparse
import json
import os
import re
import sys
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Generator, Tuple
from collections import defaultdict, Counter
from dataclasses import dataclass, field, asdict

# Rich terminal UI
try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn
    from rich import box
    console = Console()
    RICH_AVAILABLE = True
except ImportError:
    console = None
    RICH_AVAILABLE = False

# Streaming JSON parser
try:
    import ijson
    IJSON_AVAILABLE = True
except ImportError:
    IJSON_AVAILABLE = False
    print("⚠️  Install ijson for streaming: pip install ijson")

# Embeddings for clustering
try:
    from sentence_transformers import SentenceTransformer
    import numpy as np
    EMBEDDINGS_AVAILABLE = True
except ImportError:
    EMBEDDINGS_AVAILABLE = False

# Clustering
try:
    from sklearn.cluster import AgglomerativeClustering
    from sklearn.metrics.pairwise import cosine_similarity
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False

# Add paths
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))

# Import Ollama client
try:
    from shared.smart_detection import OllamaClient, UserConfig
except ImportError:
    try:
        from src.shared.smart_detection import OllamaClient, UserConfig
    except ImportError:
        OllamaClient = None
        UserConfig = None


# =============================================================================
# DATA STRUCTURES
# =============================================================================
@dataclass
class ExtractedIssue:
    """An issue extracted from a ticket or conversation."""
    ticket_id: int
    source: str  # 'initial' or 'conversation_{id}'
    text: str
    text_normalized: str
    fingerprint: str
    created_at: str
    vessel_name: Optional[str] = None
    company: Optional[str] = None
    
@dataclass
class IssueCluster:
    """A cluster of similar issues (deduplicated)."""
    cluster_id: int
    title: str = ""
    description: str = ""
    root_cause: str = ""
    module: str = ""
    severity: str = "medium"
    issue_count: int = 0
    ticket_ids: List[int] = field(default_factory=list)
    sample_texts: List[str] = field(default_factory=list)
    companies: List[str] = field(default_factory=list)
    vessels: List[str] = field(default_factory=list)
    first_seen: str = ""
    last_seen: str = ""
    is_actionable: bool = False
    action_recommendation: str = ""


# =============================================================================
# CONFIGURATION
# =============================================================================
class AnalyzerConfig:
    """Configuration for issue analysis."""
    
    # Streaming
    BATCH_SIZE = 100  # Process tickets in batches
    
    # New issue detection (high precision)
    MIN_ISSUE_LENGTH = 30  # Minimum chars for issue text
    MAX_ISSUE_LENGTH = 2000  # Truncate very long issues
    
    # Acknowledgment patterns to skip
    ACKNOWLEDGMENT_PATTERNS = [
        r'^thanks?\.?!?\s*$',
        r'^thank\s*(you|u)\.?!?\s*$',
        r'^(got\s+it|ok|okay|noted|understood|perfect|great)\.?!?\s*$',
        r'^(works?|working)\s*(now|fine|great)?\.?!?\s*$',
        r'^(issue\s+)?(resolved|fixed|solved|sorted)\.?!?\s*$',
        r'^please\s+close\.?!?\s*$',
        r'^cheers\.?!?\s*$',
        r'^confirmed\.?!?\s*$',
        r'^acknowledged?\.?!?\s*$',
        r'^received\.?!?\s*$',
        r'^will\s+do\.?!?\s*$',
        r'^sure\.?!?\s*$',
    ]
    
    # Phrases indicating NEW issue (high precision signals)
    NEW_ISSUE_SIGNALS = [
        r'\balso\b.*\b(issue|problem|error|not\s+working|fail)',
        r'\banother\s+(issue|problem|thing|matter)\b',
        r'\bseparate(ly)?\s+(issue|problem|matter)\b',
        r'\badditionally\b',
        r'\bfurthermore\b',
        r'\bby\s+the\s+way\b',
        r'\bbtw\b',
        r'\bunrelated\b',
        r'\bdifferent\s+(issue|problem|matter)\b',
        r'\bnew\s+(issue|problem|error)\b',
        r'\bone\s+more\s+(thing|issue|problem)\b',
    ]
    
    # Clustering
    SIMILARITY_THRESHOLD = 0.75  # Cosine similarity for same issue
    MIN_CLUSTER_SIZE = 2  # Minimum tickets for a cluster
    
    # AI sampling
    AI_SAMPLE_SIZE = 5  # Sample size per cluster for AI enrichment
    
    # Actionable thresholds
    ACTIONABLE_MIN_TICKETS = 3  # Minimum tickets for actionable
    ACTIONABLE_MIN_SEVERITY = "medium"  # Minimum severity


# Compile patterns
_ACK_PATTERNS = [re.compile(p, re.IGNORECASE) for p in AnalyzerConfig.ACKNOWLEDGMENT_PATTERNS]
_NEW_ISSUE_PATTERNS = [re.compile(p, re.IGNORECASE) for p in AnalyzerConfig.NEW_ISSUE_SIGNALS]


# =============================================================================
# TEXT PROCESSING
# =============================================================================
def clean_html(html_text: str) -> str:
    """Remove HTML tags and clean text."""
    if not html_text:
        return ""
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', ' ', html_text)
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    # Remove common email signatures
    text = re.sub(r'(warm\s+)?regards,?\s*[\w\s]*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'best,?\s*[\w\s]*$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'thanks,?\s*[\w\s]*$', '', text, flags=re.IGNORECASE)
    return text.strip()


def normalize_text(text: str) -> str:
    """Normalize text for fingerprinting."""
    if not text:
        return ""
    # Lowercase
    text = text.lower()
    # Remove special characters
    text = re.sub(r'[^\w\s]', ' ', text)
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    # Remove common filler words
    stopwords = {'the', 'a', 'an', 'is', 'are', 'was', 'were', 'be', 'been', 
                 'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will',
                 'would', 'could', 'should', 'may', 'might', 'must', 'shall',
                 'to', 'of', 'in', 'for', 'on', 'with', 'at', 'by', 'from',
                 'as', 'into', 'through', 'during', 'before', 'after', 'above',
                 'below', 'between', 'under', 'again', 'further', 'then', 'once',
                 'here', 'there', 'when', 'where', 'why', 'how', 'all', 'each',
                 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor',
                 'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very',
                 'just', 'also', 'now', 'please', 'hi', 'hello', 'dear', 'sir',
                 'madam', 'team', 'support', 'good', 'day', 'morning', 'afternoon',
                 'evening', 'greetings', 'kindly', 'request', 'requesting'}
    words = [w for w in text.split() if w not in stopwords and len(w) > 2]
    return ' '.join(words)


def generate_fingerprint(text: str) -> str:
    """Generate a fingerprint hash for issue deduplication."""
    normalized = normalize_text(text)
    # Use first 500 chars for fingerprint
    normalized = normalized[:500]
    return hashlib.md5(normalized.encode()).hexdigest()[:12]


def is_acknowledgment(text: str) -> bool:
    """Check if text is just an acknowledgment (not a real issue)."""
    if not text:
        return True
    text = text.strip()
    if len(text) < 10:
        return True
    for pattern in _ACK_PATTERNS:
        if pattern.match(text):
            return True
    return False


def has_new_issue_signal(text: str) -> bool:
    """Check if text contains signals of a NEW issue (high precision)."""
    for pattern in _NEW_ISSUE_PATTERNS:
        if pattern.search(text):
            return True
    return False


def extract_vessel_name(ticket: dict) -> Optional[str]:
    """Extract vessel name from ticket."""
    # Try subject first
    subject = ticket.get('subject', '')
    # Pattern: "VESSEL NAME | " or "MV VESSEL NAME"
    match = re.search(r'(?:^|\|)\s*([A-Z][A-Za-z0-9\s\-\.]{2,30})\s*(?:\||$)', subject)
    if match:
        return match.group(1).strip()
    
    # Try custom fields
    custom = ticket.get('custom_fields', {})
    if custom:
        for key in ['cf_vessel_name', 'cf_vessel', 'vessel_name', 'vessel']:
            if key in custom and custom[key]:
                return str(custom[key])
    
    return None


def extract_company_name(ticket: dict) -> Optional[str]:
    """Extract company name from ticket."""
    company = ticket.get('company', {})
    if company and isinstance(company, dict):
        return company.get('name')
    return None


# =============================================================================
# STREAMING TICKET PARSER
# =============================================================================
class StreamingTicketParser:
    """Parse tickets from large JSON files using streaming."""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.total_tickets = 0
        self.file_size = os.path.getsize(filepath)
    
    def estimate_ticket_count(self) -> int:
        """Estimate ticket count from file size."""
        # Rough estimate: ~50KB per ticket average
        return max(1, self.file_size // 50000)
    
    def stream_tickets(self) -> Generator[dict, None, None]:
        """Stream tickets one at a time using ijson."""
        if not IJSON_AVAILABLE:
            # Fallback: load entire file (not recommended for large files)
            with open(self.filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            tickets = data if isinstance(data, list) else data.get('tickets', [])
            for ticket in tickets:
                yield ticket
            return
        
        with open(self.filepath, 'rb') as f:
            # Try different JSON structures
            try:
                # Structure: [ticket, ticket, ...]
                parser = ijson.items(f, 'item')
                for ticket in parser:
                    self.total_tickets += 1
                    yield ticket
            except ijson.JSONError:
                f.seek(0)
                try:
                    # Structure: {"tickets": [ticket, ...]}
                    parser = ijson.items(f, 'tickets.item')
                    for ticket in parser:
                        self.total_tickets += 1
                        yield ticket
                except ijson.JSONError:
                    # Fallback to full load
                    f.seek(0)
                    data = json.load(f)
                    tickets = data if isinstance(data, list) else data.get('tickets', [])
                    for ticket in tickets:
                        self.total_tickets += 1
                        yield ticket


# =============================================================================
# ISSUE EXTRACTOR
# =============================================================================
class IssueExtractor:
    """Extract issues from tickets with conversation awareness."""
    
    def __init__(self, high_precision: bool = True):
        self.high_precision = high_precision
        self.stats = {
            'tickets_processed': 0,
            'initial_issues': 0,
            'conversation_issues': 0,
            'skipped_acks': 0,
            'skipped_agent': 0,
        }
    
    def extract_from_ticket(self, ticket: dict) -> List[ExtractedIssue]:
        """Extract all issues from a ticket."""
        issues = []
        ticket_id = ticket.get('id', 0)
        vessel = extract_vessel_name(ticket)
        company = extract_company_name(ticket)
        
        # 1. Initial issue (subject + description)
        initial_text = self._get_initial_issue_text(ticket)
        if initial_text and len(initial_text) >= AnalyzerConfig.MIN_ISSUE_LENGTH:
            issues.append(ExtractedIssue(
                ticket_id=ticket_id,
                source='initial',
                text=initial_text[:AnalyzerConfig.MAX_ISSUE_LENGTH],
                text_normalized=normalize_text(initial_text),
                fingerprint=generate_fingerprint(initial_text),
                created_at=ticket.get('created_at', ''),
                vessel_name=vessel,
                company=company,
            ))
            self.stats['initial_issues'] += 1
        
        # 2. Scan conversations for NEW issues
        conversations = ticket.get('conversations', [])
        if isinstance(conversations, list):
            for conv in conversations:
                conv_issue = self._extract_conversation_issue(
                    conv, ticket_id, vessel, company, issues
                )
                if conv_issue:
                    issues.append(conv_issue)
                    self.stats['conversation_issues'] += 1
        
        self.stats['tickets_processed'] += 1
        return issues
    
    def _get_initial_issue_text(self, ticket: dict) -> str:
        """Get the initial issue text from subject + description."""
        subject = ticket.get('subject', '')
        
        # Try description_text first (plain text), then description (HTML)
        description = ticket.get('description_text', '')
        if not description:
            description = clean_html(ticket.get('description', ''))
        
        # Combine subject and description
        if subject and description:
            return f"{subject}\n{description}"
        return subject or description
    
    def _extract_conversation_issue(self, conv: dict, ticket_id: int, 
                                     vessel: str, company: str,
                                     existing_issues: List[ExtractedIssue]) -> Optional[ExtractedIssue]:
        """Extract a NEW issue from a conversation message."""
        # Skip agent messages
        if not conv.get('incoming', False):
            self.stats['skipped_agent'] += 1
            return None
        
        # Get message text
        text = conv.get('body_text', '')
        if not text:
            text = clean_html(conv.get('body', ''))
        
        if not text or len(text) < AnalyzerConfig.MIN_ISSUE_LENGTH:
            return None
        
        # Skip acknowledgments
        if is_acknowledgment(text):
            self.stats['skipped_acks'] += 1
            return None
        
        # High precision: only detect if clear NEW issue signal
        if self.high_precision:
            if not has_new_issue_signal(text):
                # Check if this is substantially different from initial issue
                if existing_issues:
                    initial_fp = existing_issues[0].fingerprint
                    new_fp = generate_fingerprint(text)
                    if initial_fp == new_fp:
                        return None  # Same issue, skip
                else:
                    return None
        
        # Create issue
        conv_id = conv.get('id', 'unknown')
        return ExtractedIssue(
            ticket_id=ticket_id,
            source=f'conversation_{conv_id}',
            text=text[:AnalyzerConfig.MAX_ISSUE_LENGTH],
            text_normalized=normalize_text(text),
            fingerprint=generate_fingerprint(text),
            created_at=conv.get('created_at', ''),
            vessel_name=vessel,
            company=company,
        )


# =============================================================================
# ISSUE CLUSTERER
# =============================================================================
class IssueClustering:
    """Cluster similar issues for deduplication."""
    
    def __init__(self):
        self.model = None
        if EMBEDDINGS_AVAILABLE:
            try:
                self.model = SentenceTransformer('all-MiniLM-L6-v2')
            except Exception as e:
                print(f"⚠️ Could not load embedding model: {e}")
    
    def cluster_issues(self, issues: List[ExtractedIssue], 
                       progress_callback=None) -> List[IssueCluster]:
        """Cluster similar issues together."""
        if not issues:
            return []
        
        # Method 1: Fingerprint-based (fast, exact dedup)
        fingerprint_groups = self._group_by_fingerprint(issues)
        
        if progress_callback:
            progress_callback("Fingerprint grouping", 30)
        
        # Method 2: Embedding-based (semantic similarity)
        if self.model and SKLEARN_AVAILABLE and len(fingerprint_groups) > 1:
            clusters = self._cluster_with_embeddings(fingerprint_groups, progress_callback)
        else:
            # Fallback: use fingerprint groups as clusters
            clusters = self._fingerprint_to_clusters(fingerprint_groups)
        
        if progress_callback:
            progress_callback("Clustering complete", 60)
        
        return clusters
    
    def _group_by_fingerprint(self, issues: List[ExtractedIssue]) -> Dict[str, List[ExtractedIssue]]:
        """Group issues by exact fingerprint match."""
        groups = defaultdict(list)
        for issue in issues:
            groups[issue.fingerprint].append(issue)
        return dict(groups)
    
    def _cluster_with_embeddings(self, fingerprint_groups: Dict[str, List[ExtractedIssue]],
                                  progress_callback=None) -> List[IssueCluster]:
        """Cluster fingerprint groups using semantic embeddings."""
        # Get representative text for each fingerprint group
        group_ids = list(fingerprint_groups.keys())
        representative_texts = []
        for fp in group_ids:
            issues = fingerprint_groups[fp]
            # Use the shortest text as representative (often clearest)
            rep_text = min([i.text_normalized for i in issues], key=len)
            representative_texts.append(rep_text)
        
        if progress_callback:
            progress_callback("Generating embeddings", 40)
        
        # Generate embeddings
        embeddings = self.model.encode(representative_texts, show_progress_bar=False)
        
        if progress_callback:
            progress_callback("Computing similarities", 50)
        
        # Compute similarity matrix
        sim_matrix = cosine_similarity(embeddings)
        
        # Agglomerative clustering
        distance_matrix = 1 - sim_matrix
        np.fill_diagonal(distance_matrix, 0)
        
        clustering = AgglomerativeClustering(
            n_clusters=None,
            distance_threshold=1 - AnalyzerConfig.SIMILARITY_THRESHOLD,
            metric='precomputed',
            linkage='average'
        )
        
        labels = clustering.fit_predict(distance_matrix)
        
        # Build clusters
        cluster_groups = defaultdict(list)
        for idx, label in enumerate(labels):
            fp = group_ids[idx]
            cluster_groups[label].extend(fingerprint_groups[fp])
        
        clusters = []
        for cluster_id, issues in cluster_groups.items():
            cluster = self._build_cluster(cluster_id, issues)
            if cluster.issue_count >= AnalyzerConfig.MIN_CLUSTER_SIZE:
                clusters.append(cluster)
        
        # Sort by issue count
        clusters.sort(key=lambda c: c.issue_count, reverse=True)
        
        # Reassign cluster IDs
        for i, cluster in enumerate(clusters):
            cluster.cluster_id = i + 1
        
        return clusters
    
    def _fingerprint_to_clusters(self, fingerprint_groups: Dict[str, List[ExtractedIssue]]) -> List[IssueCluster]:
        """Convert fingerprint groups to clusters (fallback)."""
        clusters = []
        for cluster_id, (fp, issues) in enumerate(fingerprint_groups.items(), 1):
            cluster = self._build_cluster(cluster_id, issues)
            if cluster.issue_count >= AnalyzerConfig.MIN_CLUSTER_SIZE:
                clusters.append(cluster)
        
        clusters.sort(key=lambda c: c.issue_count, reverse=True)
        return clusters
    
    def _build_cluster(self, cluster_id: int, issues: List[ExtractedIssue]) -> IssueCluster:
        """Build a cluster from a list of issues."""
        ticket_ids = list(set(i.ticket_id for i in issues))
        companies = list(set(i.company for i in issues if i.company))
        vessels = list(set(i.vessel_name for i in issues if i.vessel_name))
        
        # Get sample texts (shortest ones, often clearest)
        sample_texts = sorted([i.text for i in issues], key=len)[:5]
        
        # Date range
        dates = [i.created_at for i in issues if i.created_at]
        first_seen = min(dates) if dates else ""
        last_seen = max(dates) if dates else ""
        
        return IssueCluster(
            cluster_id=cluster_id,
            issue_count=len(issues),
            ticket_ids=ticket_ids,
            sample_texts=sample_texts,
            companies=companies[:10],
            vessels=vessels[:10],
            first_seen=first_seen,
            last_seen=last_seen,
        )


# =============================================================================
# AI ENRICHMENT (with Checkpoint/Resume)
# =============================================================================
class AIEnricher:
    """Enrich clusters with AI-generated insights. Supports checkpoint/resume."""
    
    CHECKPOINT_FILE = "ai_enrichment_checkpoint.json"
    
    def __init__(self, output_dir: Path = None):
        self.ollama = OllamaClient() if OllamaClient else None
        self.available = self.ollama and self.ollama.available
        self.output_dir = output_dir or Path('reports')
        self.checkpoint_path = self.output_dir / self.CHECKPOINT_FILE
    
    def enrich_clusters(self, clusters: List[IssueCluster], 
                        progress_callback=None) -> List[IssueCluster]:
        """Enrich clusters with AI-generated titles, root causes, etc.
        
        Supports checkpoint/resume - can be interrupted and continued later.
        """
        if not self.available:
            # Fallback: generate basic titles from text
            for cluster in clusters:
                cluster.title = self._generate_basic_title(cluster)
                cluster.severity = self._estimate_severity(cluster)
                cluster.is_actionable = self._is_actionable(cluster)
            return clusters
        
        # Load existing checkpoint
        checkpoint = self._load_checkpoint()
        enriched_ids = set(checkpoint.get('enriched_cluster_ids', []))
        enriched_data = {item['cluster_id']: item for item in checkpoint.get('enriched_data', [])}
        
        total = len(clusters)
        skipped = 0
        processed = 0
        start_time = datetime.now()
        
        # Apply already-enriched data
        for cluster in clusters:
            if cluster.cluster_id in enriched_data:
                data = enriched_data[cluster.cluster_id]
                cluster.title = data.get('title', cluster.title)
                cluster.description = data.get('description', '')
                cluster.root_cause = data.get('root_cause', '')
                cluster.module = data.get('module', '')
                cluster.severity = data.get('severity', 'medium')
                cluster.action_recommendation = data.get('recommendation', '')
                cluster.is_actionable = self._is_actionable(cluster)
                skipped += 1
        
        if skipped > 0 and progress_callback:
            progress_callback(f"Resumed: {skipped} clusters already enriched", 60)
        
        # Process remaining clusters
        for i, cluster in enumerate(clusters):
            if cluster.cluster_id in enriched_ids:
                continue  # Already enriched
            
            processed += 1
            
            # Calculate ETA
            if processed > 1:
                elapsed = (datetime.now() - start_time).total_seconds()
                avg_time = elapsed / (processed - 1)
                remaining = total - skipped - processed
                eta_seconds = int(avg_time * remaining)
                eta_str = self._format_eta(eta_seconds)
            else:
                eta_str = "calculating..."
            
            if progress_callback:
                pct = 60 + int(30 * (skipped + processed) / total)
                progress_callback(
                    f"AI enriching {skipped + processed}/{total} (ETA: {eta_str})",
                    pct
                )
            
            try:
                enriched = self._enrich_single_cluster(cluster)
                cluster.title = enriched.get('title', cluster.title)
                cluster.description = enriched.get('description', '')
                cluster.root_cause = enriched.get('root_cause', '')
                cluster.module = enriched.get('module', '')
                cluster.severity = enriched.get('severity', 'medium')
                cluster.action_recommendation = enriched.get('recommendation', '')
                
                # Save to checkpoint immediately
                self._save_cluster_to_checkpoint(cluster)
                
            except Exception as e:
                # Fallback on error
                cluster.title = self._generate_basic_title(cluster)
                cluster.severity = self._estimate_severity(cluster)
            
            cluster.is_actionable = self._is_actionable(cluster)
        
        # Clear checkpoint on successful completion
        if processed > 0:
            self._finalize_checkpoint()
        
        return clusters
    
    def _load_checkpoint(self) -> Dict:
        """Load checkpoint file if exists."""
        if self.checkpoint_path.exists():
            try:
                with open(self.checkpoint_path, 'r') as f:
                    checkpoint = json.load(f)
                    print(f"📂 Loaded checkpoint: {len(checkpoint.get('enriched_cluster_ids', []))} clusters already enriched")
                    return checkpoint
            except Exception as e:
                print(f"⚠️ Could not load checkpoint: {e}")
        return {'enriched_cluster_ids': [], 'enriched_data': []}
    
    def _save_cluster_to_checkpoint(self, cluster: IssueCluster):
        """Save a single cluster to checkpoint file."""
        checkpoint = self._load_checkpoint() if self.checkpoint_path.exists() else {
            'enriched_cluster_ids': [],
            'enriched_data': [],
            'started_at': datetime.now().isoformat(),
        }
        
        # Add this cluster
        if cluster.cluster_id not in checkpoint['enriched_cluster_ids']:
            checkpoint['enriched_cluster_ids'].append(cluster.cluster_id)
            checkpoint['enriched_data'].append({
                'cluster_id': cluster.cluster_id,
                'title': cluster.title,
                'description': cluster.description,
                'root_cause': cluster.root_cause,
                'module': cluster.module,
                'severity': cluster.severity,
                'recommendation': cluster.action_recommendation,
            })
        
        checkpoint['last_updated'] = datetime.now().isoformat()
        checkpoint['total_enriched'] = len(checkpoint['enriched_cluster_ids'])
        
        # Save atomically
        temp_path = self.checkpoint_path.with_suffix('.tmp')
        with open(temp_path, 'w') as f:
            json.dump(checkpoint, f, indent=2)
        temp_path.rename(self.checkpoint_path)
    
    def _finalize_checkpoint(self):
        """Mark checkpoint as complete."""
        if self.checkpoint_path.exists():
            checkpoint = self._load_checkpoint()
            checkpoint['completed'] = True
            checkpoint['completed_at'] = datetime.now().isoformat()
            with open(self.checkpoint_path, 'w') as f:
                json.dump(checkpoint, f, indent=2)
    
    def _format_eta(self, seconds: int) -> str:
        """Format seconds as human-readable ETA."""
        if seconds < 60:
            return f"{seconds}s"
        elif seconds < 3600:
            return f"{seconds // 60}m {seconds % 60}s"
        else:
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            return f"{hours}h {minutes}m"
    
    def _enrich_single_cluster(self, cluster: IssueCluster) -> Dict:
        """Use AI to enrich a single cluster."""
        # Take sample texts
        samples = cluster.sample_texts[:AnalyzerConfig.AI_SAMPLE_SIZE]
        sample_text = "\n---\n".join(samples)
        
        prompt = f"""Analyze these customer support issues (all describing the same underlying problem):

{sample_text}

Provide a JSON response with:
1. "title": A clear, concise title (max 10 words) describing the issue
2. "description": One sentence describing the problem
3. "root_cause": Most likely root cause
4. "module": Which product module is affected (e.g., Sync, Signature, ORB, Deck Log, or General)
5. "severity": "high", "medium", or "low" based on impact
6. "recommendation": One actionable recommendation to fix this

Respond ONLY with valid JSON, no other text."""

        response = self.ollama.generate(prompt)
        
        # Parse JSON response
        try:
            # Find JSON in response
            json_match = re.search(r'\{[^{}]*\}', response, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
        except:
            pass
        
        return {}
    
    def _generate_basic_title(self, cluster: IssueCluster) -> str:
        """Generate a basic title from cluster text."""
        if not cluster.sample_texts:
            return f"Issue Cluster {cluster.cluster_id}"
        
        # Use first sample, truncate
        text = cluster.sample_texts[0]
        # Get first sentence or first 50 chars
        first_sentence = re.split(r'[.!?\n]', text)[0]
        if len(first_sentence) > 60:
            first_sentence = first_sentence[:57] + "..."
        return first_sentence
    
    def _estimate_severity(self, cluster: IssueCluster) -> str:
        """Estimate severity based on frequency and keywords."""
        # High frequency = higher severity
        if cluster.issue_count >= 20:
            return "high"
        elif cluster.issue_count >= 10:
            return "medium"
        
        # Check for severity keywords
        text = ' '.join(cluster.sample_texts).lower()
        high_keywords = ['critical', 'urgent', 'emergency', 'down', 'blocked', 'cannot', 'failure', 'crash']
        if any(kw in text for kw in high_keywords):
            return "high"
        
        return "medium"
    
    def _is_actionable(self, cluster: IssueCluster) -> bool:
        """Determine if cluster is actionable."""
        if cluster.issue_count < AnalyzerConfig.ACTIONABLE_MIN_TICKETS:
            return False
        if cluster.severity == "low":
            return False
        return True


# =============================================================================
# REPORT GENERATOR (Professional Excel with Charts)
# =============================================================================

# Excel/Charts imports
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class Colors:
    """Professional color palette."""
    PRIMARY = "1F4E79"
    PRIMARY_LIGHT = "2E75B6"
    HEADER_BG = "1F4E79"
    HEADER_TEXT = "FFFFFF"
    ROW_ALT = "F2F2F2"
    HIGH_CONF = "C6EFCE"
    HIGH_CONF_TEXT = "006100"
    MED_CONF = "FFEB9C"
    MED_CONF_TEXT = "9C5700"
    LOW_CONF = "FFC7CE"
    LOW_CONF_TEXT = "9C0006"
    CHART_COLORS = ["5B9BD5", "ED7D31", "A5A5A5", "FFC000", "4472C4", "70AD47"]


class IssueReportGenerator:
    """Generate professional reports from issue clusters."""
    
    def __init__(self, clusters: List[IssueCluster], output_dir: Path, 
                 stats: Dict = None):
        self.clusters = clusters
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.generated_at = datetime.now().strftime('%Y-%m-%d %H:%M')
        self.stats = stats or {}
    
    def generate_all(self) -> Dict[str, str]:
        """Generate all report types."""
        outputs = {}
        
        # Excel Report (main deliverable)
        if EXCEL_AVAILABLE:
            excel_path = self.generate_excel_report()
            outputs['Excel Report'] = str(excel_path)
        
        # A) Executive Summary (Top 50)
        exec_path = self.generate_executive_summary()
        outputs['Executive Summary'] = str(exec_path)
        
        # B) Complete Issue Registry
        registry_path = self.generate_complete_registry()
        outputs['Issue Registry'] = str(registry_path)
        
        # C) Actionable Issues
        actionable_path = self.generate_actionable_report()
        outputs['Actionable Issues'] = str(actionable_path)
        
        # JSON export
        json_path = self.generate_json()
        outputs['Raw Data'] = str(json_path)
        
        return outputs
    
    # =========================================================================
    # PROFESSIONAL EXCEL REPORT WITH CHARTS
    # =========================================================================
    def generate_excel_report(self) -> Path:
        """Generate professional Excel report with charts."""
        excel_path = self.output_dir / 'issue_analysis_report.xlsx'
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color=Colors.HEADER_TEXT, size=11)
        header_fill = PatternFill(start_color=Colors.HEADER_BG, end_color=Colors.HEADER_BG, fill_type="solid")
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alt_fill = PatternFill(start_color=Colors.ROW_ALT, end_color=Colors.ROW_ALT, fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Severity fills
        high_fill = PatternFill(start_color=Colors.LOW_CONF, end_color=Colors.LOW_CONF, fill_type="solid")
        med_fill = PatternFill(start_color=Colors.MED_CONF, end_color=Colors.MED_CONF, fill_type="solid")
        low_fill = PatternFill(start_color=Colors.HIGH_CONF, end_color=Colors.HIGH_CONF, fill_type="solid")
        
        # =====================================================================
        # SHEET 1: Executive Dashboard
        # =====================================================================
        ws = wb.active
        ws.title = "Executive Dashboard"
        
        # Title
        ws['A1'] = "FTEX Issue Analysis Report"
        ws['A1'].font = Font(bold=True, size=20, color=Colors.PRIMARY)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Generated: {self.generated_at}"
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:F2')
        
        # Key Metrics
        ws['A4'] = "KEY METRICS"
        ws['A4'].font = Font(bold=True, size=14, color=Colors.PRIMARY)
        
        total_tickets = sum(len(c.ticket_ids) for c in self.clusters)
        actionable = [c for c in self.clusters if c.is_actionable]
        high_sev = [c for c in self.clusters if c.severity == 'high']
        
        metrics = [
            ("Total Distinct Issues", len(self.clusters)),
            ("Total Tickets Covered", total_tickets),
            ("Actionable Issues", len(actionable)),
            ("High Severity Issues", len(high_sev)),
            ("", ""),
            ("Tickets Processed", self.stats.get('tickets_processed', 'N/A')),
            ("Initial Issues Found", self.stats.get('initial_issues', 'N/A')),
            ("Conversation Issues Found", self.stats.get('conversation_issues', 'N/A')),
        ]
        
        for i, (label, value) in enumerate(metrics, start=5):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True) if label else Font()
            ws[f'B{i}'] = value
            ws[f'B{i}'].alignment = Alignment(horizontal='right')
            if label == "High Severity Issues" and isinstance(value, int) and value > 0:
                ws[f'B{i}'].fill = high_fill
                ws[f'B{i}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 15
        
        # Top 10 Issues mini-table
        ws['A16'] = "TOP 10 ISSUES"
        ws['A16'].font = Font(bold=True, size=14, color=Colors.PRIMARY)
        
        headers = ['#', 'Issue', 'Tickets', 'Severity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=17, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        for i, cluster in enumerate(self.clusters[:10], 1):
            row = 17 + i
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=(cluster.title or f"Issue #{cluster.cluster_id}")[:50]).border = thin_border
            ws.cell(row=row, column=3, value=len(cluster.ticket_ids)).border = thin_border
            
            sev_cell = ws.cell(row=row, column=4, value=cluster.severity.title())
            sev_cell.border = thin_border
            sev_cell.alignment = Alignment(horizontal='center')
            if cluster.severity == 'high':
                sev_cell.fill = high_fill
            elif cluster.severity == 'medium':
                sev_cell.fill = med_fill
            else:
                sev_cell.fill = low_fill
            
            if row % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = alt_fill
        
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        
        # =====================================================================
        # SHEET 2: All Issues (Complete Registry)
        # =====================================================================
        ws2 = wb.create_sheet("All Issues")
        
        headers = ['#', 'Issue Title', 'Tickets', 'Severity', 'Module', 'Root Cause', 
                   'Companies', 'First Seen', 'Last Seen', 'Sample Ticket IDs']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        for i, cluster in enumerate(self.clusters, 1):
            row = i + 1
            ws2.cell(row=row, column=1, value=i).border = thin_border
            ws2.cell(row=row, column=2, value=(cluster.title or f"Issue #{cluster.cluster_id}")[:60]).border = thin_border
            ws2.cell(row=row, column=3, value=len(cluster.ticket_ids)).border = thin_border
            
            sev_cell = ws2.cell(row=row, column=4, value=cluster.severity.title())
            sev_cell.border = thin_border
            sev_cell.alignment = Alignment(horizontal='center')
            if cluster.severity == 'high':
                sev_cell.fill = high_fill
            elif cluster.severity == 'medium':
                sev_cell.fill = med_fill
            
            ws2.cell(row=row, column=5, value=cluster.module or '').border = thin_border
            ws2.cell(row=row, column=6, value=(cluster.root_cause or '')[:50]).border = thin_border
            ws2.cell(row=row, column=7, value=', '.join(cluster.companies[:3])).border = thin_border
            ws2.cell(row=row, column=8, value=cluster.first_seen[:10] if cluster.first_seen else '').border = thin_border
            ws2.cell(row=row, column=9, value=cluster.last_seen[:10] if cluster.last_seen else '').border = thin_border
            ws2.cell(row=row, column=10, value=', '.join(map(str, cluster.ticket_ids[:5]))).border = thin_border
            
            if row % 2 == 0:
                for col in range(1, 11):
                    if col != 4:  # Keep severity color
                        if not ws2.cell(row=row, column=col).fill.start_color.rgb or ws2.cell(row=row, column=col).fill.start_color.rgb == '00000000':
                            ws2.cell(row=row, column=col).fill = alt_fill
        
        ws2.auto_filter.ref = f"A1:J{len(self.clusters)+1}"
        ws2.freeze_panes = 'A2'
        
        col_widths = [5, 45, 10, 10, 12, 35, 25, 12, 12, 30]
        for i, width in enumerate(col_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = width
        
        # =====================================================================
        # SHEET 3: Actionable Issues
        # =====================================================================
        ws3 = wb.create_sheet("Actionable Issues")
        
        # Warning banner
        ws3['A1'] = "⚡ ACTION REQUIRED: Issues needing immediate attention"
        ws3['A1'].font = Font(bold=True, size=12, color=Colors.LOW_CONF_TEXT)
        ws3['A1'].fill = PatternFill(start_color=Colors.LOW_CONF, end_color=Colors.LOW_CONF, fill_type="solid")
        ws3.merge_cells('A1:G1')
        
        headers = ['Priority', 'Issue', 'Tickets', 'Severity', 'Module', 'Root Cause', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        actionable_sorted = sorted(actionable, 
                                   key=lambda c: (0 if c.severity == 'high' else 1, -len(c.ticket_ids)))
        
        for i, cluster in enumerate(actionable_sorted, 1):
            row = i + 2
            ws3.cell(row=row, column=1, value=i).border = thin_border
            ws3.cell(row=row, column=2, value=(cluster.title or f"Issue #{cluster.cluster_id}")[:50]).border = thin_border
            ws3.cell(row=row, column=3, value=len(cluster.ticket_ids)).border = thin_border
            
            sev_cell = ws3.cell(row=row, column=4, value=cluster.severity.title())
            sev_cell.border = thin_border
            sev_cell.alignment = Alignment(horizontal='center')
            sev_cell.font = Font(bold=True)
            if cluster.severity == 'high':
                sev_cell.fill = high_fill
            elif cluster.severity == 'medium':
                sev_cell.fill = med_fill
            
            ws3.cell(row=row, column=5, value=cluster.module or '').border = thin_border
            ws3.cell(row=row, column=6, value=(cluster.root_cause or '')[:40]).border = thin_border
            ws3.cell(row=row, column=7, value=(cluster.action_recommendation or '')[:50]).border = thin_border
            
            if row % 2 == 1:
                for col in range(1, 8):
                    if col != 4:
                        ws3.cell(row=row, column=col).fill = alt_fill
        
        ws3.auto_filter.ref = f"A2:G{len(actionable_sorted)+2}"
        ws3.freeze_panes = 'A3'
        
        col_widths = [8, 40, 10, 10, 12, 30, 40]
        for i, width in enumerate(col_widths, 1):
            ws3.column_dimensions[get_column_letter(i)].width = width
        
        # =====================================================================
        # SHEET 4: Charts
        # =====================================================================
        ws4 = wb.create_sheet("Charts")
        
        # Prepare data for charts
        # Top 15 issues by ticket count
        ws4['A1'] = "Top 15 Issues by Ticket Count"
        ws4['A1'].font = Font(bold=True, size=12, color=Colors.PRIMARY)
        
        ws4['A3'] = "Issue"
        ws4['B3'] = "Tickets"
        ws4['A3'].font = header_font
        ws4['B3'].font = header_font
        ws4['A3'].fill = header_fill
        ws4['B3'].fill = header_fill
        
        for i, cluster in enumerate(self.clusters[:15], 1):
            ws4.cell(row=3+i, column=1, value=(cluster.title or f"Issue {i}")[:30])
            ws4.cell(row=3+i, column=2, value=len(cluster.ticket_ids))
        
        # Bar Chart - Top Issues
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.style = 10
        chart1.title = "Top 15 Issues by Ticket Count"
        chart1.y_axis.title = "Issue"
        chart1.x_axis.title = "Number of Tickets"
        
        data = Reference(ws4, min_col=2, min_row=3, max_row=18)
        cats = Reference(ws4, min_col=1, min_row=4, max_row=18)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        chart1.width = 18
        chart1.height = 12
        
        ws4.add_chart(chart1, "D3")
        
        # Severity Distribution Data
        ws4['A22'] = "Severity Distribution"
        ws4['A22'].font = Font(bold=True, size=12, color=Colors.PRIMARY)
        
        ws4['A24'] = "Severity"
        ws4['B24'] = "Count"
        ws4['A24'].font = header_font
        ws4['B24'].font = header_font
        ws4['A24'].fill = header_fill
        ws4['B24'].fill = header_fill
        
        severity_counts = Counter(c.severity for c in self.clusters)
        for i, (sev, count) in enumerate([('high', severity_counts.get('high', 0)),
                                          ('medium', severity_counts.get('medium', 0)),
                                          ('low', severity_counts.get('low', 0))], 1):
            ws4.cell(row=24+i, column=1, value=sev.title())
            ws4.cell(row=24+i, column=2, value=count)
        
        # Pie Chart - Severity
        chart2 = PieChart()
        chart2.title = "Issues by Severity"
        
        data2 = Reference(ws4, min_col=2, min_row=24, max_row=27)
        cats2 = Reference(ws4, min_col=1, min_row=25, max_row=27)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.width = 10
        chart2.height = 8
        
        # Add data labels
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showPercent = True
        chart2.dataLabels.showVal = True
        
        ws4.add_chart(chart2, "D22")
        
        # Module Distribution (if available)
        modules = Counter(c.module for c in self.clusters if c.module)
        if modules:
            ws4['A35'] = "Issues by Module"
            ws4['A35'].font = Font(bold=True, size=12, color=Colors.PRIMARY)
            
            ws4['A37'] = "Module"
            ws4['B37'] = "Issues"
            ws4['A37'].font = header_font
            ws4['B37'].font = header_font
            ws4['A37'].fill = header_fill
            ws4['B37'].fill = header_fill
            
            for i, (module, count) in enumerate(modules.most_common(8), 1):
                ws4.cell(row=37+i, column=1, value=module[:20])
                ws4.cell(row=37+i, column=2, value=count)
            
            chart3 = BarChart()
            chart3.type = "col"
            chart3.style = 10
            chart3.title = "Issues by Module"
            
            data3 = Reference(ws4, min_col=2, min_row=37, max_row=37+len(modules.most_common(8)))
            cats3 = Reference(ws4, min_col=1, min_row=38, max_row=37+len(modules.most_common(8)))
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(cats3)
            chart3.width = 12
            chart3.height = 8
            
            ws4.add_chart(chart3, "D35")
        
        ws4.column_dimensions['A'].width = 35
        ws4.column_dimensions['B'].width = 12
        
        # =====================================================================
        # SHEET 5: Ticket Details
        # =====================================================================
        ws5 = wb.create_sheet("Ticket Details")
        
        headers = ['Issue #', 'Issue Title', 'Ticket ID', 'Company', 'Vessel']
        for col, header in enumerate(headers, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        row = 2
        for cluster in self.clusters[:100]:  # Limit to top 100 issues
            for ticket_id in cluster.ticket_ids[:20]:  # Limit tickets per issue
                ws5.cell(row=row, column=1, value=cluster.cluster_id).border = thin_border
                ws5.cell(row=row, column=2, value=(cluster.title or '')[:40]).border = thin_border
                ws5.cell(row=row, column=3, value=ticket_id).border = thin_border
                ws5.cell(row=row, column=4, value=cluster.companies[0] if cluster.companies else '').border = thin_border
                ws5.cell(row=row, column=5, value=cluster.vessels[0] if cluster.vessels else '').border = thin_border
                
                if row % 2 == 0:
                    for col in range(1, 6):
                        ws5.cell(row=row, column=col).fill = alt_fill
                row += 1
                
                if row > 5000:  # Safety limit
                    break
            if row > 5000:
                break
        
        ws5.auto_filter.ref = f"A1:E{row-1}"
        ws5.freeze_panes = 'A2'
        
        col_widths = [10, 40, 12, 25, 25]
        for i, width in enumerate(col_widths, 1):
            ws5.column_dimensions[get_column_letter(i)].width = width
        
        # Save
        wb.save(excel_path)
        return excel_path
    
    def generate_executive_summary(self) -> Path:
        """Generate top 50 issues executive summary."""
        path = self.output_dir / 'executive_summary.md'
        
        top_50 = self.clusters[:50]
        
        lines = [
            "# FTEX Issue Analysis: Executive Summary",
            "",
            f"**Generated:** {self.generated_at}",
            f"**Total Distinct Issues:** {len(self.clusters)}",
            f"**Total Tickets Analyzed:** {sum(len(c.ticket_ids) for c in self.clusters)}",
            "",
            "---",
            "",
            "## Top 50 Issues by Frequency",
            "",
        ]
        
        for i, cluster in enumerate(top_50, 1):
            severity_icon = "🔴" if cluster.severity == "high" else "🟡" if cluster.severity == "medium" else "🟢"
            lines.extend([
                f"### {i}. {cluster.title or f'Issue #{cluster.cluster_id}'}",
                "",
                f"**Tickets:** {len(cluster.ticket_ids)} | **Severity:** {severity_icon} {cluster.severity.title()}",
                "",
            ])
            if cluster.description:
                lines.append(f"{cluster.description}")
                lines.append("")
            if cluster.root_cause:
                lines.append(f"**Root Cause:** {cluster.root_cause}")
                lines.append("")
            if cluster.module:
                lines.append(f"**Module:** {cluster.module}")
                lines.append("")
            lines.append(f"**Ticket IDs:** {', '.join(map(str, cluster.ticket_ids[:10]))}" + 
                        (f"... (+{len(cluster.ticket_ids)-10} more)" if len(cluster.ticket_ids) > 10 else ""))
            lines.append("")
        
        lines.extend([
            "---",
            "*Generated by FTEX Issue Analyzer v7.0*"
        ])
        
        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return path
    
    def generate_complete_registry(self) -> Path:
        """Generate complete issue registry."""
        path = self.output_dir / 'issue_registry.json'
        
        registry = {
            'generated_at': self.generated_at,
            'total_issues': len(self.clusters),
            'total_tickets': sum(len(c.ticket_ids) for c in self.clusters),
            'issues': [asdict(c) for c in self.clusters]
        }
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(registry, f, indent=2, default=str)
        
        return path
    
    def generate_actionable_report(self) -> Path:
        """Generate actionable issues report."""
        path = self.output_dir / 'actionable_issues.md'
        
        actionable = [c for c in self.clusters if c.is_actionable]
        
        # Sort by severity then frequency
        severity_order = {'high': 0, 'medium': 1, 'low': 2}
        actionable.sort(key=lambda c: (severity_order.get(c.severity, 2), -len(c.ticket_ids)))
        
        lines = [
            "# Actionable Issues Report",
            "",
            f"**Generated:** {self.generated_at}",
            f"**Actionable Issues:** {len(actionable)} (out of {len(self.clusters)} total)",
            "",
            "---",
            "",
            "## Priority Action Items",
            "",
        ]
        
        # Group by severity
        high = [c for c in actionable if c.severity == 'high']
        medium = [c for c in actionable if c.severity == 'medium']
        
        if high:
            lines.extend(["### 🔴 High Severity", ""])
            for cluster in high:
                lines.extend(self._format_actionable_item(cluster))
        
        if medium:
            lines.extend(["### 🟡 Medium Severity", ""])
            for cluster in medium:
                lines.extend(self._format_actionable_item(cluster))
        
        lines.extend([
            "---",
            "*Generated by FTEX Issue Analyzer v7.0*"
        ])
        
        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        return path
    
    def _format_actionable_item(self, cluster: IssueCluster) -> List[str]:
        """Format a single actionable item."""
        lines = [
            f"**{cluster.title or f'Issue #{cluster.cluster_id}'}** ({len(cluster.ticket_ids)} tickets)",
            "",
        ]
        if cluster.action_recommendation:
            lines.append(f"📋 **Action:** {cluster.action_recommendation}")
            lines.append("")
        if cluster.root_cause:
            lines.append(f"🔍 **Root Cause:** {cluster.root_cause}")
            lines.append("")
        if cluster.companies:
            lines.append(f"🏢 **Companies:** {', '.join(cluster.companies[:5])}")
            lines.append("")
        lines.append(f"🎫 **Tickets:** {', '.join(map(str, cluster.ticket_ids[:5]))}")
        lines.append("")
        return lines
    
    def generate_json(self) -> Path:
        """Generate JSON export."""
        path = self.output_dir / 'analysis_data.json'
        
        data = {
            'generated_at': self.generated_at,
            'summary': {
                'total_issues': len(self.clusters),
                'total_tickets': sum(len(c.ticket_ids) for c in self.clusters),
                'actionable_count': sum(1 for c in self.clusters if c.is_actionable),
                'high_severity_count': sum(1 for c in self.clusters if c.severity == 'high'),
            },
            'clusters': [asdict(c) for c in self.clusters],
        }
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
        
        return path


# =============================================================================
# MAIN ANALYZER
# =============================================================================
class IssueAnalyzer:
    """Main issue-centric analyzer orchestrator."""
    
    def __init__(self, input_path: str, output_dir: str = 'reports',
                 use_ai: bool = True, high_precision: bool = True,
                 sample_size: Optional[int] = None):
        self.input_path = input_path
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)  # Ensure dir exists for checkpoints
        self.use_ai = use_ai
        self.high_precision = high_precision
        self.sample_size = sample_size
        
        self.parser = StreamingTicketParser(input_path)
        self.extractor = IssueExtractor(high_precision=high_precision)
        self.clusterer = IssueClustering()
        self.enricher = AIEnricher(output_dir=self.output_dir) if use_ai else None
    
    def run(self, progress_callback=None) -> Dict[str, Any]:
        """Run the complete analysis pipeline."""
        start_time = datetime.now()
        
        # Phase 1: Stream and extract issues
        if progress_callback:
            progress_callback("Streaming tickets...", 0)
        
        issues = self._extract_all_issues(progress_callback)
        
        if not issues:
            return {'error': 'No issues extracted'}
        
        # Phase 2: Cluster issues
        if progress_callback:
            progress_callback("Clustering issues...", 30)
        
        clusters = self.clusterer.cluster_issues(issues, progress_callback)
        
        # Phase 3: AI enrichment
        if self.use_ai and self.enricher and self.enricher.available:
            if progress_callback:
                progress_callback("AI enrichment...", 60)
            clusters = self.enricher.enrich_clusters(clusters, progress_callback)
        else:
            # Basic enrichment
            for cluster in clusters:
                cluster.title = cluster.title or self._basic_title(cluster)
                cluster.severity = self.enricher._estimate_severity(cluster) if self.enricher else "medium"
                cluster.is_actionable = (
                    cluster.issue_count >= AnalyzerConfig.ACTIONABLE_MIN_TICKETS and
                    cluster.severity in ['high', 'medium']
                )
        
        # Phase 4: Generate reports
        if progress_callback:
            progress_callback("Generating reports...", 90)
        
        # Build stats before report generation
        stats = {
            'tickets_processed': self.extractor.stats['tickets_processed'],
            'initial_issues': self.extractor.stats['initial_issues'],
            'conversation_issues': self.extractor.stats['conversation_issues'],
            'total_issues_extracted': len(issues),
            'clusters_created': len(clusters),
            'actionable_issues': sum(1 for c in clusters if c.is_actionable),
        }
        
        generator = IssueReportGenerator(clusters, self.output_dir, stats=stats)
        outputs = generator.generate_all()
        
        elapsed = (datetime.now() - start_time).total_seconds()
        stats['elapsed_seconds'] = elapsed
        
        return {
            'success': True,
            'stats': stats,
            'outputs': outputs,
            'clusters': clusters,
        }
    
    def _extract_all_issues(self, progress_callback=None) -> List[ExtractedIssue]:
        """Extract issues from all tickets using streaming."""
        issues = []
        count = 0
        estimated = self.parser.estimate_ticket_count()
        
        for ticket in self.parser.stream_tickets():
            ticket_issues = self.extractor.extract_from_ticket(ticket)
            issues.extend(ticket_issues)
            count += 1
            
            if self.sample_size and count >= self.sample_size:
                break
            
            if progress_callback and count % 100 == 0:
                pct = min(25, int(25 * count / estimated))
                progress_callback(f"Extracted {count:,} tickets...", pct)
        
        return issues
    
    def _basic_title(self, cluster: IssueCluster) -> str:
        """Generate basic title for cluster."""
        if cluster.sample_texts:
            text = cluster.sample_texts[0][:60]
            return text.split('\n')[0]
        return f"Issue #{cluster.cluster_id}"


# =============================================================================
# CLI & TERMINAL UI
# =============================================================================
def print_header():
    """Print header."""
    if console:
        console.print(Panel.fit(
            "[bold cyan]FTEX Issue Analyzer[/bold cyan] v7.0\n"
            "[dim]Streaming · Conversation-Aware · Deduplication[/dim]",
            border_style="cyan"
        ))
        console.print()
    else:
        print("=" * 60)
        print("FTEX Issue Analyzer v7.0")
        print("=" * 60)


def print_results(results: Dict):
    """Print results."""
    if console:
        stats = results.get('stats', {})
        
        table = Table(show_header=False, box=box.SIMPLE)
        table.add_column("", style="cyan")
        table.add_column("", justify="right")
        
        table.add_row("Tickets Processed", f"{stats.get('tickets_processed', 0):,}")
        table.add_row("Initial Issues", f"{stats.get('initial_issues', 0):,}")
        table.add_row("Conversation Issues", f"{stats.get('conversation_issues', 0):,}")
        table.add_row("Issue Clusters", f"{stats.get('clusters_created', 0):,}")
        table.add_row("Actionable Issues", f"[green]{stats.get('actionable_issues', 0):,}[/green]")
        table.add_row("Time", f"{stats.get('elapsed_seconds', 0):.1f}s")
        
        console.print(Panel(table, title="📊 Analysis Results", border_style="green"))
        
        # Print outputs
        outputs = results.get('outputs', {})
        if outputs:
            console.print()
            console.print(Panel(
                "\n".join([f"  📄 [cyan]{name}[/cyan]: {path}" for name, path in outputs.items()]),
                title="📁 Generated Files",
                border_style="green",
            ))
    else:
        print("\nResults:")
        for key, value in results.get('stats', {}).items():
            print(f"  {key}: {value}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='FTEX Issue-Centric Analyzer v7.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 issue_analyzer.py --input tickets.json
  python3 issue_analyzer.py --input tickets.json --sample 1000
  python3 issue_analyzer.py --input tickets.json --no-ai
        """
    )
    parser.add_argument('--input', '-i', required=True, help='Input tickets.json path')
    parser.add_argument('--output', '-o', default='reports', help='Output directory')
    parser.add_argument('--sample', '-s', type=int, help='Sample N tickets (for testing)')
    parser.add_argument('--no-ai', action='store_true', help='Disable AI enrichment')
    parser.add_argument('--low-precision', action='store_true', 
                        help='Use low precision for new issue detection (more false positives)')
    
    args = parser.parse_args()
    
    print_header()
    
    # Validate input
    if not Path(args.input).exists():
        print(f"❌ File not found: {args.input}")
        sys.exit(1)
    
    # Run analyzer
    analyzer = IssueAnalyzer(
        input_path=args.input,
        output_dir=args.output,
        use_ai=not args.no_ai,
        high_precision=not args.low_precision,
        sample_size=args.sample,
    )
    
    if console:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            TimeElapsedColumn(),
            console=console
        ) as progress:
            task = progress.add_task("Analyzing...", total=100)
            
            def update_progress(msg, pct):
                progress.update(task, completed=pct, description=msg)
            
            results = analyzer.run(progress_callback=update_progress)
            progress.update(task, completed=100)
    else:
        results = analyzer.run()
    
    print_results(results)
    
    if console:
        console.print()
        console.print(Panel("[bold green]✓ Analysis complete[/bold green]", border_style="green"))


if __name__ == '__main__':
    main()